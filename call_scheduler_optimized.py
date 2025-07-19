import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import itertools
from collections import defaultdict
import threading
import queue
import re
import calendar

class DatePickerDropdown:
    """Simple and reliable date picker using listbox"""
    
    def __init__(self, parent, **kwargs):
        self.parent = parent
        self.date_var = tk.StringVar()
        self.current_date = datetime.now().date()
        
        # Create the main frame
        self.frame = ttk.Frame(parent)
        
        # Create entry and button
        self.entry = ttk.Entry(self.frame, textvariable=self.date_var, width=18, **kwargs)
        self.entry.pack(side='left', fill='x', expand=True)
        
        self.button = ttk.Button(self.frame, text="📅", width=3, command=self.show_calendar)
        self.button.pack(side='right')
        
        # Calendar popup
        self.calendar_window = None
        
        # Set default date
        self.set_date(self.current_date)
    
    def pack(self, **kwargs):
        return self.frame.pack(**kwargs)
    
    def grid(self, **kwargs):
        return self.frame.grid(**kwargs)
    
    def get_date(self):
        """Get the selected date as a date object"""
        try:
            return datetime.strptime(self.date_var.get(), '%Y-%m-%d').date()
        except ValueError:
            return None
    
    def set_date(self, date):
        """Set the date"""
        if isinstance(date, str):
            try:
                date = datetime.strptime(date, '%Y-%m-%d').date()
            except ValueError:
                return
        self.current_date = date
        self.date_var.set(date.strftime('%Y-%m-%d'))
    
    def show_calendar(self):
        """Show the calendar popup"""
        if self.calendar_window:
            self.calendar_window.destroy()
        
        # Create popup window
        self.calendar_window = tk.Toplevel(self.parent)
        self.calendar_window.title("Select Friday Date")
        self.calendar_window.geometry("300x400")
        self.calendar_window.resizable(False, False)
        self.calendar_window.transient(self.parent)
        self.calendar_window.grab_set()
        
        # Position near the button
        x = self.button.winfo_rootx()
        y = self.button.winfo_rooty() + self.button.winfo_height()
        
        # Ensure window doesn't go off screen
        screen_width = self.calendar_window.winfo_screenwidth()
        screen_height = self.calendar_window.winfo_screenheight()
        
        if x + 300 > screen_width:
            x = screen_width - 300
        if y + 400 > screen_height:
            y = self.button.winfo_rooty() - 400
        
        self.calendar_window.geometry(f"+{x}+{y}")
        
        # Create calendar
        self.create_calendar_widget()
        
        # Bind escape key to close
        self.calendar_window.bind('<Escape>', lambda e: self.calendar_window.destroy())
    
    def create_calendar_widget(self):
        """Create the calendar widget"""
        # Header frame
        header_frame = ttk.Frame(self.calendar_window)
        header_frame.pack(fill='x', padx=10, pady=10)
        
        # Navigation buttons
        ttk.Button(header_frame, text="◀", width=4, 
                  command=lambda: self.change_month(-1)).pack(side='left')
        
        self.month_label = ttk.Label(header_frame, text="", font=('TkDefaultFont', 12, 'bold'))
        self.month_label.pack(side='left', expand=True)
        
        ttk.Button(header_frame, text="▶", width=4, 
                  command=lambda: self.change_month(1)).pack(side='right')
        
        # Instructions
        ttk.Label(self.calendar_window, text="Select a Friday:", font=('TkDefaultFont', 10, 'bold')).pack(pady=5)
        
        # Listbox for dates
        listbox_frame = ttk.Frame(self.calendar_window)
        listbox_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.date_listbox = tk.Listbox(listbox_frame, font=('TkDefaultFont', 11), height=15)
        scrollbar = ttk.Scrollbar(listbox_frame, orient='vertical', command=self.date_listbox.yview)
        self.date_listbox.configure(yscrollcommand=scrollbar.set)
        
        self.date_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Bind double-click to select
        self.date_listbox.bind('<Double-Button-1>', self.select_from_listbox)
        
        # Update calendar display
        self.update_calendar_display()
    
    def change_month(self, delta):
        """Change the displayed month"""
        year = self.current_date.year
        month = self.current_date.month
        
        month += delta
        if month > 12:
            month = 1
            year += 1
        elif month < 1:
            month = 12
            year -= 1
        
        self.current_date = self.current_date.replace(year=year, month=month, day=1)
        self.update_calendar_display()
    
    def update_calendar_display(self):
        """Update the calendar display"""
        # Clear listbox
        self.date_listbox.delete(0, tk.END)
        
        # Update month label
        month_name = self.current_date.strftime('%B %Y')
        self.month_label.configure(text=month_name)
        
        # Get all Fridays in the month
        fridays = []
        current_date = self.current_date.replace(day=1)
        
        # Find first Friday of the month
        while current_date.weekday() != 4:  # Friday is 4
            current_date += timedelta(days=1)
        
        # Add all Fridays in the month
        while current_date.month == self.current_date.month:
            fridays.append(current_date)
            current_date += timedelta(days=7)
        
        # Add Fridays to listbox
        for friday in fridays:
            display_text = friday.strftime('%A, %B %d, %Y')
            self.date_listbox.insert(tk.END, display_text)
    
    def select_from_listbox(self, event):
        """Select date from listbox"""
        selection = self.date_listbox.curselection()
        if selection:
            # Get the selected Friday
            fridays = []
            current_date = self.current_date.replace(day=1)
            
            # Find first Friday of the month
            while current_date.weekday() != 4:
                current_date += timedelta(days=1)
            
            # Get all Fridays in the month
            while current_date.month == self.current_date.month:
                fridays.append(current_date)
                current_date += timedelta(days=7)
            
            # Set the selected date
            if selection[0] < len(fridays):
                self.set_date(fridays[selection[0]])
                self.calendar_window.destroy()
                self.calendar_window = None

class CallSchedulerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Call Schedule Generator")
        self.root.geometry("1000x700")
        
        # Performance optimizations
        self.root.option_add('*tearOff', False)  # Disable tear-off menus
        self.root.update_idletasks()
        
        # Database setup
        self.db_path = "call_schedule.db"
        self.init_database()
        
        # Variables
        self.start_date = tk.StringVar()
        self.end_date = tk.StringVar()
        self.excluded_weeks = []
        
        # Threading for schedule generation
        self.schedule_queue = queue.Queue()
        self.generating = False
        
        self.create_widgets()
        self.load_data()
        
        # Bind window close event
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def validate_date_format(self, date_str):
        """Validate date format and ensure it's a Friday"""
        try:
            # Check format
            if not re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
                return False, "Date must be in YYYY-MM-DD format"
            
            # Parse date
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            # Check if it's a Friday
            if date_obj.weekday() != 4:  # Friday is 4
                return False, "Date must be a Friday"
            
            return True, date_obj
        except ValueError:
            return False, "Invalid date format"
    
    def on_closing(self):
        """Handle application closing"""
        if self.generating:
            self.generating = False
        self.root.destroy()
        
    def init_database(self):
        """Initialize SQLite database with tables"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Create groups table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS groups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
        ''')
        
        # Create personnel table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS personnel (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                group_id INTEGER,
                FOREIGN KEY (group_id) REFERENCES groups (id)
            )
        ''')
        
        # Create excluded weeks table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS excluded_weeks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                week_start TEXT NOT NULL
            )
        ''')
        
        # Create personnel excluded weeks table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS personnel_excluded_weeks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                personnel_id INTEGER,
                week_start TEXT NOT NULL,
                FOREIGN KEY (personnel_id) REFERENCES personnel (id)
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def create_widgets(self):
        """Create the main GUI widgets with performance optimizations"""
        # Create notebook for tabs
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Groups tab
        self.groups_frame = ttk.Frame(notebook)
        notebook.add(self.groups_frame, text="Groups")
        self.create_groups_tab()
        
        # Personnel tab
        self.personnel_frame = ttk.Frame(notebook)
        notebook.add(self.personnel_frame, text="Personnel")
        self.create_personnel_tab()
        
        # Schedule tab
        self.schedule_frame = ttk.Frame(notebook)
        notebook.add(self.schedule_frame, text="Schedule")
        self.create_schedule_tab()
        
        # Personnel Excluded Weeks tab
        self.personnel_excluded_frame = ttk.Frame(notebook)
        notebook.add(self.personnel_excluded_frame, text="Personnel Excluded Weeks")
        self.create_personnel_excluded_weeks_tab()
    
    def create_groups_tab(self):
        """Create the groups management tab"""
        # Add group section
        add_frame = ttk.LabelFrame(self.groups_frame, text="Add Group", padding=5)
        add_frame.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(add_frame, text="Group Name:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.group_name_entry = ttk.Entry(add_frame, width=25)
        self.group_name_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Button(add_frame, text="Add Group", command=self.add_group).grid(row=0, column=2, padx=5, pady=5)
        
        # Groups list section
        list_frame = ttk.LabelFrame(self.groups_frame, text="Groups", padding=5)
        list_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Treeview for groups with performance optimizations
        columns = ('ID', 'Name')
        self.groups_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        self.groups_tree.heading('ID', text='ID')
        self.groups_tree.heading('Name', text='Group Name')
        self.groups_tree.column('ID', width=50, minwidth=50)
        self.groups_tree.column('Name', width=200, minwidth=150)
        
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.groups_tree.yview)
        self.groups_tree.configure(yscrollcommand=scrollbar.set)
        
        self.groups_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Delete button
        ttk.Button(list_frame, text="Delete Selected Group", command=self.delete_group).pack(pady=5)
    
    def create_personnel_tab(self):
        """Create the personnel management tab"""
        # Add personnel section
        add_frame = ttk.LabelFrame(self.personnel_frame, text="Add Personnel", padding=5)
        add_frame.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(add_frame, text="Name:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.personnel_name_entry = ttk.Entry(add_frame, width=25)
        self.personnel_name_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(add_frame, text="Group:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.personnel_group_var = tk.StringVar()
        self.personnel_group_combo = ttk.Combobox(add_frame, textvariable=self.personnel_group_var, width=18)
        self.personnel_group_combo.grid(row=0, column=3, padx=5, pady=5)
        
        ttk.Button(add_frame, text="Add Personnel", command=self.add_personnel).grid(row=0, column=4, padx=5, pady=5)
        
        # Personnel list section
        list_frame = ttk.LabelFrame(self.personnel_frame, text="Personnel", padding=5)
        list_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Treeview for personnel with performance optimizations
        columns = ('ID', 'Name', 'Group')
        self.personnel_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        self.personnel_tree.heading('ID', text='ID')
        self.personnel_tree.heading('Name', text='Name')
        self.personnel_tree.heading('Group', text='Group')
        self.personnel_tree.column('ID', width=50, minwidth=50)
        self.personnel_tree.column('Name', width=200, minwidth=150)
        self.personnel_tree.column('Group', width=150, minwidth=100)
        
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.personnel_tree.yview)
        self.personnel_tree.configure(yscrollcommand=scrollbar.set)
        
        self.personnel_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Buttons frame
        buttons_frame = ttk.Frame(list_frame)
        buttons_frame.pack(pady=5)
        
        ttk.Button(buttons_frame, text="Delete Selected", command=self.delete_personnel).pack(side='left', padx=5)
        ttk.Button(buttons_frame, text="Reassign Group", command=self.reassign_personnel).pack(side='left', padx=5)
    
    def create_schedule_tab(self):
        """Create the schedule generation tab"""
        # Date selection section
        date_frame = ttk.LabelFrame(self.schedule_frame, text="Schedule Period", padding=5)
        date_frame.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(date_frame, text="Start Date (Friday 7 AM):").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.start_date_picker = DatePickerDropdown(date_frame)
        self.start_date_picker.grid(row=0, column=1, padx=5, pady=5)
        self.start_date_picker.set_date("2025-01-03")  # Default Friday
        
        ttk.Label(date_frame, text="End Date:").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.end_date_picker = DatePickerDropdown(date_frame)
        self.end_date_picker.grid(row=0, column=3, padx=5, pady=5)
        self.end_date_picker.set_date("2025-01-31")  # Default end date
        
        # Generate button with progress indicator
        self.generate_button = ttk.Button(date_frame, text="Generate Schedules", command=self.generate_schedules)
        self.generate_button.grid(row=0, column=4, padx=20, pady=5)
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(date_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=1, column=0, columnspan=5, sticky='ew', padx=5, pady=5)
        
        # Results section
        results_frame = ttk.LabelFrame(self.schedule_frame, text="Generated Schedules", padding=5)
        results_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Text widget to show results with performance optimizations
        self.results_text = tk.Text(results_frame, wrap='word', height=15, font=('TkDefaultFont', 9))
        scrollbar = ttk.Scrollbar(results_frame, orient='vertical', command=self.results_text.yview)
        self.results_text.configure(yscrollcommand=scrollbar.set)
        
        self.results_text.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Export button
        ttk.Button(results_frame, text="Export to Excel", command=self.export_to_excel).pack(pady=5)
    
    def create_personnel_excluded_weeks_tab(self):
        """Create the personnel excluded weeks management tab"""
        # Add personnel excluded week section
        add_frame = ttk.LabelFrame(self.personnel_excluded_frame, text="Add Personnel Excluded Weeks", padding=5)
        add_frame.pack(fill='x', padx=5, pady=5)
        
        # Personnel selection
        ttk.Label(add_frame, text="Select Personnel:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.personnel_excluded_var = tk.StringVar()
        self.personnel_excluded_combo = ttk.Combobox(add_frame, textvariable=self.personnel_excluded_var, width=25)
        self.personnel_excluded_combo.grid(row=0, column=1, padx=5, pady=5)
        
        # Week selection
        ttk.Label(add_frame, text="Select Week (Friday):").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        self.personnel_week_picker = DatePickerDropdown(add_frame)
        self.personnel_week_picker.grid(row=0, column=3, padx=5, pady=5)
        self.personnel_week_picker.set_date("2025-01-03")  # Default Friday
        
        # Add button
        ttk.Button(add_frame, text="Add Excluded Week", command=self.add_personnel_excluded_week).grid(row=0, column=4, padx=5, pady=5)
        
        # Personnel excluded weeks list section
        list_frame = ttk.LabelFrame(self.personnel_excluded_frame, text="Personnel Excluded Weeks", padding=5)
        list_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Treeview for personnel excluded weeks
        columns = ('ID', 'Personnel', 'Week Start')
        self.personnel_excluded_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        self.personnel_excluded_tree.heading('ID', text='ID')
        self.personnel_excluded_tree.heading('Personnel', text='Personnel')
        self.personnel_excluded_tree.heading('Week Start', text='Week Start (Friday)')
        self.personnel_excluded_tree.column('ID', width=50, minwidth=50)
        self.personnel_excluded_tree.column('Personnel', width=200, minwidth=150)
        self.personnel_excluded_tree.column('Week Start', width=200, minwidth=150)
        
        scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.personnel_excluded_tree.yview)
        self.personnel_excluded_tree.configure(yscrollcommand=scrollbar.set)
        
        self.personnel_excluded_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Delete button
        ttk.Button(list_frame, text="Delete Selected", command=self.delete_personnel_excluded_week).pack(pady=5)
    
    def load_data(self):
        """Load data from database into GUI"""
        self.load_groups()
        self.load_personnel()
        self.load_personnel_excluded_weeks()
    
    def load_groups(self):
        """Load groups into the groups treeview"""
        # Clear existing items efficiently
        for item in self.groups_tree.get_children():
            self.groups_tree.delete(item)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT id, name FROM groups ORDER BY name")
        groups = cursor.fetchall()
        conn.close()
        
        # Batch insert for better performance
        for group in groups:
            self.groups_tree.insert('', 'end', values=group)
        
        # Update personnel group combo
        self.personnel_group_combo['values'] = [group[1] for group in groups]
        
        # Update personnel excluded combo
        self.personnel_excluded_combo['values'] = [group[1] for group in groups]
    
    def load_personnel(self):
        """Load personnel into the personnel treeview"""
        # Clear existing items efficiently
        for item in self.personnel_tree.get_children():
            self.personnel_tree.delete(item)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT p.id, p.name, g.name 
            FROM personnel p 
            LEFT JOIN groups g ON p.group_id = g.id 
            ORDER BY p.name
        ''')
        personnel = cursor.fetchall()
        conn.close()
        
        # Batch insert for better performance
        for person in personnel:
            self.personnel_tree.insert('', 'end', values=person)
        
        # Update personnel excluded combo with actual personnel names
        self.personnel_excluded_combo['values'] = [f"{person[1]} ({person[2]})" for person in personnel]
    

    
    def add_group(self):
        """Add a new group"""
        name = self.group_name_entry.get().strip()
        if not name:
            messagebox.showerror("Error", "Please enter a group name")
            return
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("INSERT INTO groups (name) VALUES (?)", (name,))
            conn.commit()
            conn.close()
            
            self.group_name_entry.delete(0, tk.END)
            self.load_groups()
            messagebox.showinfo("Success", f"Group '{name}' added successfully")
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "Group name already exists")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add group: {str(e)}")
    
    def delete_group(self):
        """Delete selected group"""
        selection = self.groups_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a group to delete")
            return
        
        item = self.groups_tree.item(selection[0])
        group_id = item['values'][0]
        group_name = item['values'][1]
        
        # Check if group has personnel
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM personnel WHERE group_id = ?", (group_id,))
        count = cursor.fetchone()[0]
        conn.close()
        
        if count > 0:
            messagebox.showerror("Error", f"Cannot delete group '{group_name}' - it has {count} personnel assigned")
            return
        
        if messagebox.askyesno("Confirm", f"Are you sure you want to delete group '{group_name}'?"):
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM groups WHERE id = ?", (group_id,))
                conn.commit()
                conn.close()
                
                self.load_groups()
                self.load_personnel()
                messagebox.showinfo("Success", f"Group '{group_name}' deleted successfully")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete group: {str(e)}")
    
    def add_personnel(self):
        """Add new personnel"""
        name = self.personnel_name_entry.get().strip()
        group_name = self.personnel_group_var.get()
        
        if not name:
            messagebox.showerror("Error", "Please enter a name")
            return
        
        if not group_name:
            messagebox.showerror("Error", "Please select a group")
            return
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Get group ID
            cursor.execute("SELECT id FROM groups WHERE name = ?", (group_name,))
            group_result = cursor.fetchone()
            if not group_result:
                messagebox.showerror("Error", "Selected group not found")
                return
            
            group_id = group_result[0]
            
            cursor.execute("INSERT INTO personnel (name, group_id) VALUES (?, ?)", (name, group_id))
            conn.commit()
            conn.close()
            
            self.personnel_name_entry.delete(0, tk.END)
            self.personnel_group_var.set('')
            self.load_personnel()
            messagebox.showinfo("Success", f"Personnel '{name}' added successfully")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add personnel: {str(e)}")
    
    def delete_personnel(self):
        """Delete selected personnel"""
        selection = self.personnel_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select personnel to delete")
            return
        
        item = self.personnel_tree.item(selection[0])
        person_id = item['values'][0]
        person_name = item['values'][1]
        
        if messagebox.askyesno("Confirm", f"Are you sure you want to delete '{person_name}'?"):
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM personnel WHERE id = ?", (person_id,))
                conn.commit()
                conn.close()
                
                self.load_personnel()
                messagebox.showinfo("Success", f"Personnel '{person_name}' deleted successfully")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete personnel: {str(e)}")
    
    def reassign_personnel(self):
        """Reassign personnel to different group"""
        selection = self.personnel_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select personnel to reassign")
            return
        
        item = self.personnel_tree.item(selection[0])
        person_id = item['values'][0]
        person_name = item['values'][1]
        
        # Create dialog for group selection
        dialog = tk.Toplevel(self.root)
        dialog.title("Reassign Personnel")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()
        
        ttk.Label(dialog, text=f"Reassign '{person_name}' to:").pack(pady=10)
        
        group_var = tk.StringVar()
        group_combo = ttk.Combobox(dialog, textvariable=group_var, width=20)
        group_combo.pack(pady=10)
        
        # Load groups
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM groups ORDER BY name")
        groups = [row[0] for row in cursor.fetchall()]
        conn.close()
        
        group_combo['values'] = groups
        
        def confirm_reassign():
            new_group = group_var.get()
            if not new_group:
                messagebox.showerror("Error", "Please select a group", parent=dialog)
                return
            
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM groups WHERE name = ?", (new_group,))
                group_id = cursor.fetchone()[0]
                cursor.execute("UPDATE personnel SET group_id = ? WHERE id = ?", (group_id, person_id))
                conn.commit()
                conn.close()
                
                self.load_personnel()
                dialog.destroy()
                messagebox.showinfo("Success", f"'{person_name}' reassigned to '{new_group}'")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to reassign: {str(e)}", parent=dialog)
        
        ttk.Button(dialog, text="Confirm", command=confirm_reassign).pack(pady=10)
        ttk.Button(dialog, text="Cancel", command=dialog.destroy).pack(pady=5)
    
    def load_personnel_excluded_weeks(self):
        """Load personnel excluded weeks into the treeview"""
        # Clear existing items efficiently
        for item in self.personnel_excluded_tree.get_children():
            self.personnel_excluded_tree.delete(item)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT pew.id, p.name, pew.week_start 
            FROM personnel_excluded_weeks pew
            JOIN personnel p ON pew.personnel_id = p.id
            ORDER BY p.name, pew.week_start
        ''')
        personnel_weeks = cursor.fetchall()
        conn.close()
        
        # Batch insert for better performance
        for week in personnel_weeks:
            self.personnel_excluded_tree.insert('', 'end', values=week)
    
    def add_personnel_excluded_week(self):
        """Add excluded week for specific personnel"""
        personnel_name = self.personnel_excluded_var.get()
        week_date = self.personnel_week_picker.get_date()
        
        if not personnel_name:
            messagebox.showerror("Error", "Please select personnel")
            return
        
        if not week_date:
            messagebox.showerror("Error", "Please select a valid week date")
            return
        
        week_str = week_date.strftime('%Y-%m-%d')
        
        # Extract personnel name from "Name (Group)" format
        personnel_name_clean = personnel_name.split(' (')[0]
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Get personnel ID
            cursor.execute("SELECT id FROM personnel WHERE name = ?", (personnel_name_clean,))
            personnel_result = cursor.fetchone()
            if not personnel_result:
                messagebox.showerror("Error", "Selected personnel not found")
                return
            
            personnel_id = personnel_result[0]
            
            # Check if already exists
            cursor.execute("SELECT id FROM personnel_excluded_weeks WHERE personnel_id = ? AND week_start = ?", 
                         (personnel_id, week_str))
            if cursor.fetchone():
                messagebox.showerror("Error", "This week is already excluded for this personnel")
                return
            
            cursor.execute("INSERT INTO personnel_excluded_weeks (personnel_id, week_start) VALUES (?, ?)", 
                         (personnel_id, week_str))
            conn.commit()
            conn.close()
            
            self.load_personnel_excluded_weeks()
            messagebox.showinfo("Success", f"Excluded week starting {week_str} added for {personnel_name_clean}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add excluded week: {str(e)}")
    
    def delete_personnel_excluded_week(self):
        """Delete selected personnel excluded week"""
        selection = self.personnel_excluded_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select an entry to delete")
            return
        
        item = self.personnel_excluded_tree.item(selection[0])
        week_id = item['values'][0]
        personnel_name = item['values'][1]
        week_start = item['values'][2]
        
        if messagebox.askyesno("Confirm", f"Are you sure you want to remove excluded week starting {week_start} for {personnel_name}?"):
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM personnel_excluded_weeks WHERE id = ?", (week_id,))
                conn.commit()
                conn.close()
                
                self.load_personnel_excluded_weeks()
                messagebox.showinfo("Success", "Excluded week removed successfully")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to remove excluded week: {str(e)}")
    
    def generate_schedules(self):
        """Generate call schedules based on rules"""
        if self.generating:
            messagebox.showwarning("Warning", "Schedule generation already in progress")
            return
        
        # Get start date
        start_date = self.start_date_picker.get_date()
        if not start_date:
            messagebox.showerror("Error", "Please select a valid start date")
            return
        
        # Get end date
        end_date = self.end_date_picker.get_date()
        if not end_date:
            messagebox.showerror("Error", "Please select a valid end date")
            return
        
        if start_date >= end_date:
            messagebox.showerror("Error", "End date must be after start date")
            return
        
        # Get personnel and groups
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT p.id, p.name, g.name 
            FROM personnel p 
            LEFT JOIN groups g ON p.group_id = g.id 
            ORDER BY p.name
        ''')
        personnel = cursor.fetchall()
        
        # Get personnel excluded weeks
        cursor.execute('''
            SELECT p.name, pew.week_start 
            FROM personnel_excluded_weeks pew
            JOIN personnel p ON pew.personnel_id = p.id
        ''')
        personnel_excluded_weeks = {}
        for row in cursor.fetchall():
            personnel_name, week_start = row
            if personnel_name not in personnel_excluded_weeks:
                personnel_excluded_weeks[personnel_name] = []
            personnel_excluded_weeks[personnel_name].append(week_start)
        
        conn.close()
        
        if not personnel:
            messagebox.showerror("Error", "No personnel found. Please add personnel first.")
            return
        
        # Start generation in background thread
        self.generating = True
        self.generate_button.config(state='disabled')
        self.progress_var.set(0)
        
        # Start background thread
        thread = threading.Thread(target=self.generate_schedules_thread, 
                                args=(personnel, personnel_excluded_weeks, start_date, end_date))
        thread.daemon = True
        thread.start()
        
        # Start progress monitoring
        self.root.after(100, self.check_generation_progress)
    
    def generate_schedules_thread(self, personnel, personnel_excluded_weeks, start_date, end_date):
        """Generate schedules in background thread"""
        try:
            # Generate weeks
            weeks = []
            current_date = start_date
            while current_date <= end_date:
                week_str = current_date.strftime('%Y-%m-%d')
                weeks.append(week_str)
                current_date += timedelta(days=7)
            
            if not weeks:
                self.schedule_queue.put(("error", "No available weeks in the selected period"))
                return
            
            # Generate schedules
            schedules = self.generate_valid_schedules(personnel, weeks, personnel_excluded_weeks)
            
            # Put results in queue
            self.schedule_queue.put(("success", (schedules, weeks)))
            
        except Exception as e:
            self.schedule_queue.put(("error", str(e)))
    
    def check_generation_progress(self):
        """Check for generation completion"""
        try:
            result_type, result_data = self.schedule_queue.get_nowait()
            
            if result_type == "success":
                schedules, weeks = result_data
                self.display_schedules(schedules, weeks)
            else:
                messagebox.showerror("Error", result_data)
            
            self.generating = False
            self.generate_button.config(state='normal')
            self.progress_var.set(100)
            
        except queue.Empty:
            # Still generating, update progress and check again
            current_progress = self.progress_var.get()
            if current_progress < 90:
                self.progress_var.set(current_progress + 5)
            self.root.after(200, self.check_generation_progress)
    
    def generate_valid_schedules(self, personnel, weeks, personnel_excluded_weeks):
        """Generate valid schedules based on rules"""
        schedules = []
        max_schedules = 10
        
        # Group personnel by group
        groups = defaultdict(list)
        for person in personnel:
            person_id, name, group_name = person
            groups[group_name].append((person_id, name))
        
        # Create all possible assignments
        all_personnel = [(p[0], p[1], p[2]) for p in personnel]
        
        # Try different combinations
        attempts = 0
        max_attempts = 5000  # Reduced for better performance
        
        while len(schedules) < max_schedules and attempts < max_attempts:
            attempts += 1
            
            # Shuffle personnel for variety
            import random
            random.shuffle(all_personnel)
            
            schedule = self.try_create_schedule(all_personnel, weeks, groups, personnel_excluded_weeks)
            if schedule:
                # Check if this schedule is unique
                schedule_tuple = tuple(tuple(week) for week in schedule)
                if schedule_tuple not in [tuple(tuple(week) for week in s) for s in schedules]:
                    schedules.append(schedule)
        
        return schedules
    
    def try_create_schedule(self, personnel, weeks, groups, personnel_excluded_weeks):
        """Try to create a valid schedule"""
        schedule = []
        used_personnel = set()
        last_group_week = {}  # Track last week each group was used
        
        for week_idx, week in enumerate(weeks):
            week_assignments = []
            
            # Find available personnel for this week
            available = []
            for person_id, name, group_name in personnel:
                if person_id in used_personnel:
                    continue
                
                # Check if group was used in last 2 weeks
                if group_name in last_group_week:
                    if week_idx - last_group_week[group_name] < 2:
                        continue
                
                # Check if this person has excluded this week
                if name in personnel_excluded_weeks and week in personnel_excluded_weeks[name]:
                    continue
                
                available.append((person_id, name, group_name))
            
            if not available:
                # No available personnel, mark as blank
                week_assignments.append("BLANK")
            else:
                # Select personnel for this week
                for person_id, name, group_name in available:
                    if person_id not in used_personnel:
                        week_assignments.append(name)
                        used_personnel.add(person_id)
                        last_group_week[group_name] = week_idx
                        break
            
            schedule.append(week_assignments)
        
        return schedule if len(schedule) == len(weeks) else None
    
    def display_schedules(self, schedules, weeks):
        """Display generated schedules in the results text widget"""
        self.results_text.delete(1.0, tk.END)
        
        if not schedules:
            self.results_text.insert(tk.END, "No valid schedules found.\n")
            self.results_text.insert(tk.END, "Try adjusting the personnel, groups, or time period.\n")
            return
        
        self.results_text.insert(tk.END, f"Generated {len(schedules)} valid schedule(s):\n\n")
        
        for i, schedule in enumerate(schedules, 1):
            self.results_text.insert(tk.END, f"Schedule Option {i}:\n")
            self.results_text.insert(tk.END, "-" * 50 + "\n")
            
            for j, week in enumerate(weeks):
                week_date = datetime.strptime(week, '%Y-%m-%d')
                week_display = week_date.strftime('%Y-%m-%d')
                
                if j < len(schedule) and schedule[j]:
                    if schedule[j][0] == "BLANK":
                        self.results_text.insert(tk.END, f"Week {j+1} ({week_display}): BLANK\n")
                    else:
                        self.results_text.insert(tk.END, f"Week {j+1} ({week_display}): {', '.join(schedule[j])}\n")
                else:
                    self.results_text.insert(tk.END, f"Week {j+1} ({week_display}): BLANK\n")
            
            self.results_text.insert(tk.END, "\n")
        
        # Store schedules for export
        self.current_schedules = schedules
        self.current_weeks = weeks
    
    def export_to_excel(self):
        """Export schedules to Excel file"""
        if not hasattr(self, 'current_schedules') or not self.current_schedules:
            messagebox.showwarning("Warning", "No schedules to export. Please generate schedules first.")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Schedule as Excel"
        )
        
        if not filename:
            return
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Call Schedules"
            
            # Headers
            ws['A1'] = "Week"
            ws['A2'] = "Week Start (Friday)"
            
            # Add schedule columns
            for i, schedule in enumerate(self.current_schedules, 1):
                col_letter = openpyxl.utils.get_column_letter(i + 1)
                ws[f'{col_letter}1'] = f"Option {i}"
            
            # Add week data
            for i, week in enumerate(self.current_weeks, 3):
                week_date = datetime.strptime(week, '%Y-%m-%d')
                week_display = week_date.strftime('%Y-%m-%d')
                
                ws[f'A{i}'] = f"Week {i-2}"
                ws[f'B{i}'] = week_display
                
                # Add schedule data
                for j, schedule in enumerate(self.current_schedules):
                    col_letter = openpyxl.utils.get_column_letter(j + 2)
                    if i-3 < len(schedule) and schedule[i-3]:
                        if schedule[i-3][0] == "BLANK":
                            ws[f'{col_letter}{i}'] = "BLANK"
                            # Highlight blank cells
                            ws[f'{col_letter}{i}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        else:
                            ws[f'{col_letter}{i}'] = ', '.join(schedule[i-3])
                    else:
                        ws[f'{col_letter}{i}'] = "BLANK"
                        ws[f'{col_letter}{i}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Format headers
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
            for cell in ws[2]:
                cell.font = header_font
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            messagebox.showinfo("Success", f"Schedule exported to {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to Excel: {str(e)}")

def main():
    root = tk.Tk()
    app = CallSchedulerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main() 